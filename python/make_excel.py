import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = "/Users/shahbazshaikh/Sofia_Ahmed/ontario-property-tax-validation/data/property_tax_validation.db"

BLUE_DARK  = "1F4E79"
BLUE_MID   = "2E75B6"
BLUE_LIGHT = "D6E4F0"
GREEN_BG   = "C6EFCE"
GREEN_FONT = "276221"
YELLOW_BG  = "FFEB9C"
YELLOW_FNT = "9C5700"
RED_BG     = "FFC7CE"
RED_FONT   = "9C0006"
GREY_BG    = "F2F2F2"
WHITE      = "FFFFFF"

def hdr(cell, bg=BLUE_DARK, fc=WHITE, bold=True, sz=11):
    cell.font = Font(name='Arial', bold=bold, color=fc, size=sz)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def subhdr(cell):
    cell.font = Font(name='Arial', bold=True, color=WHITE, size=10)
    cell.fill = PatternFill('solid', start_color=BLUE_MID)
    cell.alignment = Alignment(horizontal='center', vertical='center')

def bdr(cell):
    s = Side(style='thin', color='BFBFBF')
    cell.border = Border(left=s, right=s, top=s, bottom=s)

def flag_fmt(cell, flag):
    if flag == 'GREEN':
        cell.fill = PatternFill('solid', start_color=GREEN_BG)
        cell.font = Font(name='Arial', bold=True, color=GREEN_FONT, size=10)
    elif flag == 'YELLOW':
        cell.fill = PatternFill('solid', start_color=YELLOW_BG)
        cell.font = Font(name='Arial', bold=True, color=YELLOW_FNT, size=10)
    elif flag in ('RED', 'MISSING', 'CRITICAL', 'HIGH', 'MEDIUM'):
        cell.fill = PatternFill('solid', start_color=RED_BG)
        cell.font = Font(name='Arial', bold=True, color=RED_FONT, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')

def alt_row(ws, row, ncols, start_col=1):
    if row % 2 == 0:
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=row, column=c)
            if cell.fill.patternType != 'solid' or cell.fill.fgColor.rgb in ('00000000','FFFFFFFF','FF000000'):
                cell.fill = PatternFill('solid', start_color=BLUE_LIGHT)

conn = sqlite3.connect(DB_PATH)
cur  = conn.cursor()
wb   = openpyxl.Workbook()

# ============================================================
# SHEET 1 -- Cover
# ============================================================
ws1 = wb.active
ws1.title = "Cover"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 3
ws1.column_dimensions['B'].width = 25
ws1.column_dimensions['C'].width = 75

ws1.merge_cells('B2:H5')
c = ws1['B2']
c.value = "Ontario Property Tax\nData Quality & Validation Framework"
c.font = Font(name='Arial', bold=True, size=18, color=WHITE)
c.fill = PatternFill('solid', start_color=BLUE_DARK)
c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws1.row_dimensions[2].height = 25
ws1.row_dimensions[3].height = 25
ws1.row_dimensions[4].height = 25
ws1.row_dimensions[5].height = 25

items = [
    ("Data Source",    "Ontario MMAH -- Financial Information Return (FIR) portal"),
    ("Coverage",       "10 Ontario municipalities -- 2022 and 2023"),
    ("Key Finding",    "Hamilton did not submit FIR for 2023 (deadline: May 31 2024)"),
    ("Methodology",    "Implied rate = Tax Levied / CVA -- validated against year over year trends"),
    ("Purpose",        "Validate that reported property tax levies are consistent with assessed values"),
]
row = 7
for label, val in items:
    ws1[f'B{row}'] = label
    ws1[f'B{row}'].font = Font(name='Arial', bold=True, size=11, color=BLUE_DARK)
    ws1[f'C{row}'] = val
    ws1[f'C{row}'].font = Font(name='Arial', size=11)
    ws1[f'C{row}'].alignment = Alignment(wrap_text=True)
    ws1.row_dimensions[row].height = 25
    row += 2

# ============================================================
# SHEET 2 -- Validation Calculator
# ============================================================
ws2 = wb.create_sheet("Validation Calculator")
ws2.sheet_view.showGridLines = False

ws2.merge_cells('A1:J1')
ws2['A1'] = "Property Tax Validation -- Expected vs Reported Levy (Residential Class)"
hdr(ws2['A1'], sz=13)
ws2.row_dimensions[1].height = 35

ws2.merge_cells('A2:J2')
ws2['A2'] = "Expected Tax = CVA x Implied Rate  |  Variance = Reported - Expected  |  GREEN < 5%  |  YELLOW 5-10%  |  RED > 10%"
ws2['A2'].font = Font(name='Arial', size=9, italic=True, color="595959")
ws2['A2'].fill = PatternFill('solid', start_color=GREY_BG)
ws2['A2'].alignment = Alignment(horizontal='center')
ws2.row_dimensions[2].height = 18

hdrs2 = ['Municipality','Year','Property Class','CVA ($)','Implied Rate (%)','Expected Tax ($)','Reported Tax ($)','Variance ($)','Variance (%)','Flag']
for i, h in enumerate(hdrs2, 1):
    c = ws2.cell(row=3, column=i, value=h)
    subhdr(c); bdr(c)
ws2.row_dimensions[3].height = 30

rows2 = cur.execute("""
    SELECT municipality, year, property_class,
           cva, implied_rate_pct, total_tax_levied,
           CASE WHEN 0=1 THEN 'MISSING' ELSE 'GREEN' END as flag_level
    FROM silver_fir_tax_levy
    WHERE slc_code = '0010'
    ORDER BY year, municipality
""").fetchall()

dr = 4
for muni, year, pc, cva, rate, reported, flag in rows2:
    ws2.cell(row=dr, column=1, value=muni)
    ws2.cell(row=dr, column=2, value=year)
    ws2.cell(row=dr, column=3, value=pc)

    c4 = ws2.cell(row=dr, column=4, value=cva or 0)
    c4.number_format = '#,##0'

    c5 = ws2.cell(row=dr, column=5, value=round((rate or 0)/100, 6))
    c5.number_format = '0.0000%'

    c6 = ws2.cell(row=dr, column=6)
    c6.value = f'=D{dr}*E{dr}'
    c6.number_format = '#,##0'

    c7 = ws2.cell(row=dr, column=7, value=reported or 0)
    c7.number_format = '#,##0'

    c8 = ws2.cell(row=dr, column=8)
    c8.value = f'=G{dr}-F{dr}'
    c8.number_format = '#,##0'

    c9 = ws2.cell(row=dr, column=9)
    c9.value = f'=IFERROR(H{dr}/F{dr},0)'
    c9.number_format = '0.00%'

    c10 = ws2.cell(row=dr, column=10, value=flag or 'GREEN')
    flag_fmt(c10, flag or 'GREEN')

    for col in range(1, 10):
        cell = ws2.cell(row=dr, column=col)
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='right' if col > 3 else 'left', vertical='center')
        bdr(cell)
    alt_row(ws2, dr, 9)
    dr += 1

widths2 = [18, 7, 18, 22, 18, 22, 22, 18, 14, 12]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A4'

# ============================================================
# SHEET 3 -- YoY Comparison
# ============================================================
ws3 = wb.create_sheet("YoY Comparison")
ws3.sheet_view.showGridLines = False

ws3.merge_cells('A1:H1')
ws3['A1'] = "Year over Year Levy Comparison -- 2022 vs 2023 (All Property Classes)"
hdr(ws3['A1'], sz=13)
ws3.row_dimensions[1].height = 35

ws3.merge_cells('A2:H2')
ws3['A2'] = "Flag: GREEN < 10%  |  YELLOW 10-20%  |  RED > 20%  |  MISSING = no 2023 submission"
ws3['A2'].font = Font(name='Arial', size=9, italic=True, color="595959")
ws3['A2'].fill = PatternFill('solid', start_color=GREY_BG)
ws3['A2'].alignment = Alignment(horizontal='center')

hdrs3 = ['Municipality','Property Class','Levy 2022 ($M)','Levy 2023 ($M)','Change ($M)','Change (%)','CVA Change (%)','Flag']
for i, h in enumerate(hdrs3, 1):
    c = ws3.cell(row=3, column=i, value=h)
    subhdr(c); bdr(c)
ws3.row_dimensions[3].height = 30

rows3 = cur.execute("""
    SELECT municipality, property_class,
           levy_2022, levy_2023, levy_change_pct, cva_change_pct, flag_level
    FROM gold_yoy_comparison
    WHERE slc_code = '0010'
    ORDER BY flag_level DESC, ABS(COALESCE(levy_change_pct,0)) DESC
""").fetchall()

dr3 = 4
for muni, pc, l22, l23, chg, cvachg, flag in rows3:
    ws3.cell(row=dr3, column=1, value=muni)
    ws3.cell(row=dr3, column=2, value=pc)

    c3 = ws3.cell(row=dr3, column=3, value=round(l22/1e6,2) if l22 else None)
    if l22: c3.number_format = '#,##0.00'

    c4 = ws3.cell(row=dr3, column=4, value=round(l23/1e6,2) if l23 else 'MISSING')
    if l23: c4.number_format = '#,##0.00'

    c5 = ws3.cell(row=dr3, column=5)
    c5.value = f'=IFERROR(D{dr3}-C{dr3},"-")'
    c5.number_format = '#,##0.00'

    c6 = ws3.cell(row=dr3, column=6, value=round(chg/100,4) if chg else None)
    if chg: c6.number_format = '0.00%'

    c7 = ws3.cell(row=dr3, column=7, value=round(cvachg/100,4) if cvachg else None)
    if cvachg: c7.number_format = '0.00%'

    c8 = ws3.cell(row=dr3, column=8, value=flag)
    flag_fmt(c8, flag)

    for col in range(1, 8):
        cell = ws3.cell(row=dr3, column=col)
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='right' if col > 2 else 'left', vertical='center')
        bdr(cell)
    alt_row(ws3, dr3, 7)
    dr3 += 1

widths3 = [18, 20, 18, 18, 16, 14, 16, 12]
for i, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = 'A4'

# ============================================================
# SHEET 4 -- Municipality Scorecard
# ============================================================
ws4 = wb.create_sheet("Municipality Scorecard")
ws4.sheet_view.showGridLines = False

ws4.merge_cells('A1:G1')
ws4['A1'] = "Municipality Data Quality Scorecard"
hdr(ws4['A1'], sz=13)
ws4.row_dimensions[1].height = 35

ws4.merge_cells('A2:G2')
ws4['A2'] = "Submission (40pts) + Data Quality (40pts) + YoY Stability (20pts) = 100pts  |  GREEN >= 80  |  YELLOW >= 50  |  RED < 50"
ws4['A2'].font = Font(name='Arial', size=9, italic=True, color="595959")
ws4['A2'].fill = PatternFill('solid', start_color=GREY_BG)
ws4['A2'].alignment = Alignment(horizontal='center')

hdrs4 = ['Municipality','Year','Submission (40)','Data Quality (40)','YoY Stability (20)','Total Score','Flag']
for i, h in enumerate(hdrs4, 1):
    c = ws4.cell(row=3, column=i, value=h)
    subhdr(c); bdr(c)
ws4.row_dimensions[3].height = 30

rows4 = cur.execute("""
    SELECT municipality, year, submission_score,
           data_quality_score, yoy_stability_score, overall_score, overall_flag
    FROM gold_municipality_scorecard
    ORDER BY year, overall_score DESC
""").fetchall()

dr4 = 4
for muni, year, sub, dq, yoy, total, flag in rows4:
    ws4.cell(row=dr4, column=1, value=muni)
    ws4.cell(row=dr4, column=2, value=year)
    ws4.cell(row=dr4, column=3, value=sub)
    ws4.cell(row=dr4, column=4, value=dq)
    ws4.cell(row=dr4, column=5, value=yoy)

    c6 = ws4.cell(row=dr4, column=6)
    c6.value = f'=C{dr4}+D{dr4}+E{dr4}'
    c6.font = Font(name='Arial', bold=True, size=11)
    c6.alignment = Alignment(horizontal='center', vertical='center')
    bdr(c6)

    c7 = ws4.cell(row=dr4, column=7, value=flag)
    flag_fmt(c7, flag)
    bdr(c7)

    for col in range(1, 6):
        cell = ws4.cell(row=dr4, column=col)
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='center' if col > 1 else 'left', vertical='center')
        bdr(cell)
    alt_row(ws4, dr4, 5)
    dr4 += 1

widths4 = [18, 8, 18, 20, 22, 14, 12]
for i, w in enumerate(widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = 'A4'

# ============================================================
# SHEET 5 -- Anomaly Summary
# ============================================================
ws5 = wb.create_sheet("Anomaly Summary")
ws5.sheet_view.showGridLines = False

ws5.merge_cells('A1:F1')
ws5['A1'] = "Data Quality Anomalies -- Action Required"
hdr(ws5['A1'], sz=13)
ws5.row_dimensions[1].height = 35

hdrs5 = ['Municipality','Year','Anomaly Type','Description','Severity','Status']
for i, h in enumerate(hdrs5, 1):
    c = ws5.cell(row=2, column=i, value=h)
    subhdr(c); bdr(c)
ws5.row_dimensions[2].height = 30

rows5 = cur.execute("""
    SELECT municipality, year, anomaly_type,
           description, severity, status
    FROM gold_anomaly_log
    ORDER BY CASE severity WHEN 'CRITICAL' THEN 1 WHEN 'HIGH' THEN 2 WHEN 'MEDIUM' THEN 3 ELSE 4 END
""").fetchall()

dr5 = 3
for muni, year, atype, desc, sev, status in rows5:
    ws5.cell(row=dr5, column=1, value=muni)
    ws5.cell(row=dr5, column=2, value=year)
    ws5.cell(row=dr5, column=3, value=atype)
    dc = ws5.cell(row=dr5, column=4, value=desc)
    dc.alignment = Alignment(wrap_text=True, vertical='center')
    dc.font = Font(name='Arial', size=10)
    bdr(dc)

    sc = ws5.cell(row=dr5, column=5, value=sev)
    flag_fmt(sc, sev)
    bdr(sc)

    stc = ws5.cell(row=dr5, column=6, value=status)
    stc.font = Font(name='Arial', size=10)
    stc.alignment = Alignment(horizontal='center', vertical='center')
    bdr(stc)

    for col in [1, 2, 3]:
        cell = ws5.cell(row=dr5, column=col)
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='center' if col > 1 else 'left', vertical='center')
        bdr(cell)

    ws5.row_dimensions[dr5].height = 50
    dr5 += 1

widths5 = [18, 8, 25, 65, 12, 12]
for i, w in enumerate(widths5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 6 -- Rate Comparison
# ============================================================
ws6 = wb.create_sheet("Rate Comparison")
ws6.sheet_view.showGridLines = False

ws6.merge_cells('A1:E1')
ws6['A1'] = "Residential Implied Tax Rate -- 2022 vs 2023 (sorted by highest rate)"
hdr(ws6['A1'], sz=13)
ws6.row_dimensions[1].height = 35

ws6.merge_cells('A2:E2')
ws6['A2'] = "Implied Rate = Total Tax Levied / CVA x 100  |  All rates from FIR Schedule 26A"
ws6['A2'].font = Font(name='Arial', size=9, italic=True, color="595959")
ws6['A2'].fill = PatternFill('solid', start_color=GREY_BG)
ws6['A2'].alignment = Alignment(horizontal='center')

hdrs6 = ['Municipality','Rate 2022 (%)','Rate 2023 (%)','Change (%)','Note']
for i, h in enumerate(hdrs6, 1):
    c = ws6.cell(row=3, column=i, value=h)
    subhdr(c); bdr(c)
ws6.row_dimensions[3].height = 30

rows6 = cur.execute("""
    SELECT s22.municipality,
           ROUND(s22.implied_rate_pct, 4),
           ROUND(s23.implied_rate_pct, 4)
    FROM silver_fir_tax_levy s22
    LEFT JOIN silver_fir_tax_levy s23
        ON s22.municipality = s23.municipality
        AND s23.year = 2023
        AND s23.slc_code = '0010'
    WHERE s22.year = 2022 AND s22.slc_code = '0010'
    ORDER BY s23.implied_rate_pct DESC
""").fetchall()

dr6 = 4
for muni, r22, r23 in rows6:
    ws6.cell(row=dr6, column=1, value=muni)

    c2 = ws6.cell(row=dr6, column=2, value=(r22 or 0)/100)
    c2.number_format = '0.0000%'

    c3 = ws6.cell(row=dr6, column=3, value=(r23 or 0)/100 if r23 else None)
    if r23: c3.number_format = '0.0000%'

    c4 = ws6.cell(row=dr6, column=4)
    c4.value = f'=IFERROR(C{dr6}-B{dr6},"-")'
    c4.number_format = '0.0000%'

    note = ""
    if r23 and r23 > 1.5:
        note = "High rate -- above Ontario average"
    elif muni == "Hamilton":
        note = "2023 submission missing"
    ws6.cell(row=dr6, column=5, value=note)

    for col in range(1, 6):
        cell = ws6.cell(row=dr6, column=col)
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='right' if col > 1 else 'left', vertical='center')
        bdr(cell)
    alt_row(ws6, dr6, 5)
    dr6 += 1

widths6 = [18, 20, 20, 16, 35]
for i, w in enumerate(widths6, 1):
    ws6.column_dimensions[get_column_letter(i)].width = w
ws6.freeze_panes = 'A4'

OUTPUT = "/Users/shahbazshaikh/Sofia_Ahmed/ontario-property-tax-validation/data/property_tax_validation.xlsx"
wb.save(OUTPUT)
conn.close()
print("Done:", OUTPUT)