import os
import openpyxl
import pandas as pd


# These are the property class codes used in Schedule 26A
# Each code maps to a type of property in Ontario
PROPERTY_CLASSES = {
    '0010': 'Residential',
    '0050': 'Multi-residential',
    '0110': 'Farmland',
    '0140': 'Managed Forest',
    '0210': 'Commercial',
    '0510': 'Industrial',
    '0610': 'Large Industrial',
    '0710': 'Pipelines',
    '9170': 'Supplementary Taxes',
    '9180': 'Total Levied by Rate',
}


def get_municipality_name(filename):
    # File names have spaces: FI220614 Ottawa C.xlsx
    # We want just the middle part: Ottawa
    name = filename.replace('.xlsx', '').strip()
    parts = name.split(' ')
    # First part is FI220614, last part is C
    # Everything in between is the municipality name
    municipality = ' '.join(parts[1:-1])
    return municipality


def extract_data_from_file(filepath, municipality, year):
    # This function opens one FIR Excel file and reads Schedule 26A
    # Schedule 26A has the tax levy breakdown by property class

    wb = openpyxl.load_workbook(filepath, read_only=True)

    # Make sure this sheet exists
    if '26A' not in wb.sheetnames:
        print(f"    No 26A sheet found for {municipality} {year}")
        return []

    sheet = wb['26A']
    rows_data = []

    # Go through each row in the sheet
    for row in sheet.iter_rows(max_row=150, values_only=True):

        # Skip blank rows
        if not any(cell is not None for cell in row):
            continue

        # Column index 2 has the SLC code (property class code)
        slc_code = row[2]
        if slc_code is None:
            continue

        slc_code = str(slc_code).strip().zfill(4)

        # Only keep rows that match our property class codes
        if slc_code not in PROPERTY_CLASSES:
            continue

        # Pull values from specific columns
        # These column positions were identified by looking at the actual Excel files
        cva             = row[7]   # Current Value Assessment
        total_tax       = row[12]  # Total tax levied
        lt_municipal    = row[14]  # Lower tier / single tier municipal tax
        ut_municipal    = row[15]  # Upper tier municipal tax (regional municipalities)
        education       = row[16]  # Education tax (set by province)

        # Convert to float if it's a number, otherwise leave as None
        def to_float(value):
            if isinstance(value, (int, float)):
                return float(value)
            return None

        row_dict = {
            'municipality':     municipality,
            'year':             year,
            'slc_code':         slc_code,
            'property_class':   PROPERTY_CLASSES[slc_code],
            'cva':              to_float(cva),
            'total_tax_levied': to_float(total_tax),
            'lt_municipal_tax': to_float(lt_municipal),
            'ut_municipal_tax': to_float(ut_municipal),
            'education_tax':    to_float(education),
        }

        rows_data.append(row_dict)

    return rows_data


def process_folder(year):
    # Process all Excel files inside the 2022 or 2023 folder
    folder_path = f"data/raw/{year}"
    all_rows = []

    # Get list of files in folder
    all_files = os.listdir(folder_path)

    # Keep only proper xlsx files -- skip hidden files like .DS_Store
    excel_files = []
    for f in all_files:
        if f.endswith('.xlsx') and not f.startswith('.'):
            excel_files.append(f)

    excel_files = sorted(excel_files)

    print(f"\nYear {year} -- found {len(excel_files)} Excel files")

    for filename in excel_files:
        filepath = os.path.join(folder_path, filename)
        municipality = get_municipality_name(filename)

        try:
            rows = extract_data_from_file(filepath, municipality, year)
            all_rows.extend(rows)
            print(f"  ✓ {municipality:<25} {len(rows)} rows")
        except Exception as e:
            print(f"  ✗ {municipality:<25} SKIPPED -- {e}")

    return all_rows


def main():
    # Process 2022 folder
    rows_2022 = process_folder(2022)
    df_2022 = pd.DataFrame(rows_2022)
    df_2022.to_csv('data/processed/fir_data_2022.csv', index=False)
    print(f"\nSaved fir_data_2022.csv -- {len(df_2022)} rows")

    # Process 2023 folder
    rows_2023 = process_folder(2023)
    df_2023 = pd.DataFrame(rows_2023)
    df_2023.to_csv('data/processed/fir_data_2023.csv', index=False)
    print(f"Saved fir_data_2023.csv -- {len(df_2023)} rows")


main()